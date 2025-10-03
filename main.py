import psycopg2
import pandas as pd
import matplotlib.pyplot as plt
import os
import plotly.express as px
from tabulate import tabulate
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import ColorScaleRule

# Данные подключения
DB_NAME = "olist_db"
DB_USER = "postgres"
DB_PASS = "0000"
DB_HOST = "localhost"
DB_PORT = "5432"

# Папки для экспорта
CHARTS_DIR = "charts"
EXPORTS_DIR = "exports"
os.makedirs(CHARTS_DIR, exist_ok=True)
os.makedirs(EXPORTS_DIR, exist_ok=True)

# Нормальные имена листов для Excel
query_names = {
    1: "Customers Preview",
    2: "Delivered Orders",
    3: "Orders by Status",
    4: "Top Payments",
    5: "Avg Order Value by Month",
    6: "Top Sellers",
    7: "Customers by State",
    8: "Product Ratings",
    9: "Avg Delivery Time",
    10: "Top Products",
    11: "Top Customers",
    12: "Freight by Seller State",
    13: "Orders Summary"
}

def visualize_charts(conn):
    """6 графиков для аналитики"""
    # 1. Pie chart: количество заказов по статусам
    query = """
    SELECT order_status, COUNT(*) AS total_orders
    FROM olist_orders_dataset
    GROUP BY order_status;
    """
    df = pd.read_sql(query, conn)
    df.set_index("order_status")["total_orders"].plot.pie(
        autopct="%1.1f%%", figsize=(6, 6), title="Распределение заказов по статусам"
    )
    plt.ylabel("")
    plt.savefig(os.path.join(CHARTS_DIR, "pie_orders_status.png"))
    plt.close()

    # 2. Bar chart: средний чек по месяцам
    query = """
    SELECT DATE_TRUNC('month', o.order_purchase_timestamp) AS month,
           AVG(p.payment_value) AS avg_order_value
    FROM olist_orders_dataset o
    JOIN olist_order_payments_dataset p ON o.order_id = p.order_id
    GROUP BY month
    ORDER BY month;
    """
    df = pd.read_sql(query, conn)
    df.plot.bar(x="month", y="avg_order_value", title="Средний чек по месяцам", legend=False)
    plt.xlabel("Месяц")
    plt.ylabel("Средний чек")
    plt.xticks(rotation=45)
    plt.tight_layout()
    plt.savefig(os.path.join(CHARTS_DIR, "bar_avg_order_value.png"))
    plt.close()

    # 3. Horizontal bar: топ продавцы по продажам
    query = """
    SELECT s.seller_name, SUM(oi.price) AS total_sales
    FROM olist_order_items_dataset oi
    JOIN olist_sellers_dataset s ON oi.seller_id = s.seller_id
    GROUP BY s.seller_name
    ORDER BY total_sales DESC
    LIMIT 10;
    """
    df = pd.read_sql(query, conn)
    df.plot.barh(x="seller_name", y="total_sales", title="Топ-10 продавцов по продажам", legend=False)
    plt.xlabel("Сумма продаж")
    plt.ylabel("Продавец")
    plt.tight_layout()
    plt.savefig(os.path.join(CHARTS_DIR, "barh_top_sellers.png"))
    plt.close()

    # 4. Line chart: динамика количества заказов
    query = """
    SELECT DATE_TRUNC('month', order_purchase_timestamp) AS month,
           COUNT(*) AS total_orders
    FROM olist_orders_dataset
    GROUP BY month
    ORDER BY month;
    """
    df = pd.read_sql(query, conn)
    df.plot.line(x="month", y="total_orders", title="Динамика количества заказов")
    plt.xlabel("Месяц")
    plt.ylabel("Количество заказов")
    plt.savefig(os.path.join(CHARTS_DIR, "line_orders_over_time.png"))
    plt.close()

    # 5. Histogram: распределение суммы платежей
    query = """
    SELECT payment_value
    FROM olist_order_payments_dataset;
    """
    df = pd.read_sql(query, conn)
    df["payment_value"].plot.hist(bins=30, title="Распределение суммы платежей")
    plt.xlabel("Сумма платежа")
    plt.ylabel("Количество")
    plt.savefig(os.path.join(CHARTS_DIR, "hist_payment_value.png"))
    plt.close()

    # 6. Scatter: сумма покупок клиентов
    query = """
    SELECT c.customer_name, SUM(p.payment_value) AS total_spent, COUNT(o.order_id) AS total_orders
    FROM olist_orders_dataset o
    JOIN olist_order_payments_dataset p ON o.order_id = p.order_id
    JOIN olist_customers_dataset c ON o.customer_id = c.customer_id
    GROUP BY c.customer_name
    ORDER BY total_spent DESC
    LIMIT 50;
    """
    df = pd.read_sql(query, conn)
    df.plot.scatter(x="total_orders", y="total_spent", title="Траты клиентов vs. Количество заказов")
    plt.xlabel("Количество заказов")
    plt.ylabel("Общая сумма")
    plt.savefig(os.path.join(CHARTS_DIR, "scatter_customers_spent.png"))
    plt.close()

    print("✅ 6 графиков сохранены в /charts/")

def plot_time_slider(conn):
    """Интерактивный график с ползунком времени (Plotly)"""
    query = """
    SELECT DATE_TRUNC('month', o.order_purchase_timestamp) AS month,
           c.customer_state,
           AVG(p.payment_value) AS avg_order_value
    FROM olist_orders_dataset o
    JOIN olist_order_payments_dataset p ON o.order_id = p.order_id
    JOIN olist_customers_dataset c ON o.customer_id = c.customer_id
    GROUP BY month, c.customer_state
    ORDER BY month;
    """
    df = pd.read_sql(query, conn)
    df["month"] = df["month"].dt.strftime("%Y-%m")

    fig = px.scatter(
        df,
        x="customer_state",
        y="avg_order_value",
        animation_frame="month",
        size="avg_order_value",
        color="customer_state",
        title="Динамика среднего чека по месяцам и штатам"
    )
    fig.show()

def export_to_excel(dataframes_dict, filename):
    """Экспорт в Excel с форматированием"""
    filepath = os.path.join(EXPORTS_DIR, filename)
    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        for sheet_name, df in dataframes_dict.items():
            df.to_excel(writer, sheet_name=sheet_name[:31], index=False)

    wb = load_workbook(filepath)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.freeze_panes = "B2"
        ws.auto_filter.ref = ws.dimensions
        for col in ws.iter_cols(min_row=2, max_row=ws.max_row, min_col=2, max_col=ws.max_column):
            col_letter = col[0].column_letter
            range_str = f"{col_letter}2:{col_letter}{ws.max_row}"
            rule = ColorScaleRule(
                start_type="min", start_color="FFAA0000",
                mid_type="percentile", mid_value=50, mid_color="FFFFFF00",
                end_type="max", end_color="FF00AA00"
            )
            ws.conditional_formatting.add(range_str, rule)
    wb.save(filepath)

    total_rows = sum(df.shape[0] for df in dataframes_dict.values())
    print(f"✅ Created file {filename}, {len(dataframes_dict)} sheets, {total_rows} rows")

def main():
    try:
        conn = psycopg2.connect(
            dbname=DB_NAME, user=DB_USER, password=DB_PASS, host=DB_HOST, port=DB_PORT
        )
        cursor = conn.cursor()
        print("✅ Connected to database!")

        with open("queries.sql", "r") as f:
            sql_script = f.read()
        queries = [q.strip() for q in sql_script.split(";") if q.strip()]

        dataframes_dict = {}
        for i, query in enumerate(queries, start=1):
            try:
                df = pd.read_sql(query, conn)
                sheet_name = query_names.get(i, f"Query_{i}")[:31]
                dataframes_dict[sheet_name] = df

                print(f"\n--- Executing Query {i} ({sheet_name}) ---")
                print(f"📊 Shape: {df.shape[0]} rows × {df.shape[1]} cols")

                if len(df) > 10:
                    preview = pd.concat([df.head(5), df.tail(5)])
                else:
                    preview = df

                print(tabulate(preview, headers='keys', tablefmt='psql', showindex=False))

            except Exception as e:
                print(f"⚠️ Error in query {i}: {e}")

        visualize_charts(conn)
        export_to_excel(dataframes_dict, "olist_report.xlsx")
        plot_time_slider(conn)

        cursor.close()
        conn.close()
        print("\n✅ Done!")

    except Exception as e:
        print("❌ Connection error:", e)

if __name__ == "__main__":
    main()
